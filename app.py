import streamlit as st
import pdfplumber
import re
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlwt
import tempfile
import io

st.set_page_config(page_title="💧 水費自動處理系統", layout="wide")
st.title("💧 水費自動處理系統")

# ── 水號對照 ──────────────────────────────────────────────────
def load_lookup(file):
    xls = pd.read_excel(file, sheet_name=None, header=0, engine='openpyxl', dtype=str)
    cat_map, lookup = {}, {}
    for name, df in xls.items():
        if '分類' in name:
            for _, row in df.iterrows():
                c = str(row.iloc[0]).strip(); v = str(row.iloc[2]).strip()
                if c not in ('nan','') and v not in ('nan',''):
                    try: cat_map[str(int(float(c))).zfill(3)] = v
                    except: cat_map[c] = v
        else:
            cols = list(df.columns)
            for i in range(0, len(cols)-1, 3):
                wc, sc = cols[i], cols[i+1]
                for _, row in df.iterrows():
                    w = str(row[wc]).strip(); s = str(row[sc]).strip()
                    if w not in ('nan','') and s not in ('nan',''):
                        if re.match(r'^\d+\.0$', w): w = w[:-2]
                        if re.match(r'^\d+\.0$', s): s = s[:-2]
                        lookup[w] = s
    return lookup, cat_map

def parse_store(raw, cat_map):
    raw = str(raw).strip()
    m = re.match(r'^(\d+)[-\s]*(.*)', raw)
    if m:
        code = str(int(m.group(1))).zfill(3)
        label = m.group(2).strip() or code
        cat = cat_map.get(code, '門市')
    else:
        code = raw; label = raw
        cat = cat_map.get(raw, '總公司')
    return code, cat, label

# ── PDF 解析 ──────────────────────────────────────────────────
def parse_beishui(file_bytes, filename):
    wn_m = (re.search(r'北水[_\(]([A-Z0-9Y0-9-]+)[_\)]\d', filename) or
        re.search(r'北水([A-Z0-9Y][A-Z0-9-]+)\D*\d{6}', filename))
    water_no = wn_m.group(1).replace('-','') if wn_m else ''
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
    def find(pats):
        for pat in pats:
            m = re.search(pat, text, re.DOTALL)
            if m: return m.group(1).strip()
        return ''
    pay_date = find([r'(\d{3}/\d{2}/\d{2})\s+\$\d'])
    period   = find([r'用水計費期間[：:]\s*(\d{7}/\d{7})'])
    invoice  = find([r'本期發票號碼[：:]\s*(\w+)\('])
    tax_m  = re.search(r'營業稅[：:]\s*(\d+\.?\d*)\)', text)
    tax    = int(float(tax_m.group(1))) if tax_m else 0
    tot_m = (re.search(r'應繳總金額（元）Total Amount Due：\s*(\d+\.?\d*)', text) or
         re.search(r'Total Amount Due：\s*(\d+\.?\d*)', text) or
         re.search(r'應繳總金額.*?：\s*(\d+\.?\d*)', text) or
         re.search(r'\$(\d+)', text))
    total  = int(float(tot_m.group(1))) if tot_m else 0
    addr_m = (re.search(r'用水地址：(.+)', text) or
          re.search(r'Address of Water Consumption[)）]\s*\n(.+)', text))
    address = addr_m.group(1).strip() if addr_m else ''
    name_m = re.search(r'用戶姓名[：:](.+?)(?:\n|\(Customer)', text, re.DOTALL)
    name = name_m.group(1).replace('\n','').strip() if name_m else '大成鋼隆美家居室內裝修設計股份有限公司'
    bc_m = re.search(r'(\d{5})(BB[A-Z0-9]{8})\n', text)
    inv_ym  = bc_m.group(1) if bc_m else ''
    carrier = bc_m.group(2) if bc_m else ''
    return dict(water_no=water_no, pay_date=pay_date, period=period,
                invoice=invoice, inv_ym=inv_ym, carrier=carrier,
                tax_id='03774909', tax=tax, fee=total-tax, total=total,
                name=name, address=address)

def parse_taishui(file_bytes, filename):
    wn_m = re.search(r'台水([A-Z0-9]+)[_\(]', filename)
    water_no = wn_m.group(1) if wn_m else ''
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        text = '\n'.join(p.extract_text() or '' for p in pdf.pages)
    def find(pats):
        for pat in pats:
            m = re.search(pat, text, re.DOTALL)
            if m: return m.group(1).strip()
        return ''
    pay_date = find([r'本期水費扣繳日\s*\n?\s*(\d{3}/\d{2}/\d{2})'])
    pm = re.search(r'(\d{3}/\d{2}/\d{2})\s*-\s*(\d{3}/\d{2}/\d{2})', text)
    period  = f"{pm.group(1)}-{pm.group(2)}" if pm else ''
    invoice = find([r'(AN\d+)'])
    tax_id  = find([r'本公司營利事業統一編號\s+(\d{8})'])
    tax_m   = re.search(r'營業稅\s+(\d+\.?\d*)元', text)
    tax     = int(float(tax_m.group(1))) if tax_m else 0
    tot_m   = re.search(r'代繳\(代收\)總金額\s+(\d+)元', text)
    total   = int(tot_m.group(1)) if tot_m else 0
    addr_m  = re.search(r'用水地址\n(.+?)(?:\n單據號碼)', text, re.DOTALL)
    address = addr_m.group(1).replace('\n','').strip() if addr_m else ''
    bc_m    = re.search(r'(\d{5})\s+(BB[A-Z0-9]{8})', text)
    inv_ym  = bc_m.group(1) if bc_m else ''
    carrier = bc_m.group(2) if bc_m else ''
    return dict(water_no=water_no, pay_date=pay_date, period=period,
                invoice=invoice, inv_ym=inv_ym, carrier=carrier,
                tax_id=tax_id, tax=tax, fee=total-tax, total=total,
                name='大成鋼隆美家居室內裝修設計股份有限公司',
                address=address)

def process_pdf(file_bytes, filename):
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ''
        # 先用內文判斷
        if '臺北自來水事業處' in text:
            return parse_beishui(file_bytes, filename)
        elif '台灣自來水' in text:
            return parse_taishui(file_bytes, filename)
        # 內文判斷失敗時用檔名判斷
        elif '北水' in filename:
            return parse_beishui(file_bytes, filename)
        elif '台水' in filename:
            return parse_taishui(file_bytes, filename)
        else:
            st.warning(f"無法辨識（內文）：{filename}\n前100字：{text[:100]}")
            return None
    except Exception as e:
        st.error(f"讀取失敗 {filename}：{e}")
        return None

# ── 產出 Excel ────────────────────────────────────────────────
def bdr(c="AAAAAA"):
    t=Side(style="thin",color=c); return Border(left=t,right=t,top=t,bottom=t)
def cx(h="center"): return Alignment(horizontal=h,vertical="center")
def fl(c): return PatternFill("solid",start_color=c)
FH=Font(name="Arial",bold=True,color="FFFFFF",size=11)
FD=Font(name="Arial",size=11); FT=Font(name="Arial",bold=True,size=11)
FC=Font(name="Arial",bold=True,size=11,color="FFFFFF")
BDR=bdr(); BDR2=bdr("888888")
HDR=["門市代號","分類","水號","繳費日期","用水計費期間","發票號碼","發票年月",
     "載具流水號","水費公司統一編號","營業稅","水費","應繳總金額","用戶姓名","用水地址"]
NCOL=len(HDR); COL_W=[10,8,16,12,22,16,10,14,16,9,10,12,44,36]

def wcell(ws,r,c,val=None,font=None,f=None,align=None,b=None,fmt=None):
    cl=ws.cell(row=r,column=c,value=val)
    if font: cl.font=font
    if f:    cl.fill=f
    if align:cl.alignment=align
    if b:    cl.border=b
    if fmt:  cl.number_format=fmt
    return cl

def col_align(c):
    if c<=9: return cx()
    if c<=12: return cx("right")
    return cx("left")

def make_xlsx(records):
    wb=Workbook(); ws=wb.active; ws.title="工作表1"
    for i,w in enumerate(COL_W,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=22
    for c,h in enumerate(HDR,1): wcell(ws,1,c,h,FH,fl("4472C4"),cx(),BDR)
    row=2; grand=0
    for cat_name,cf in [("門市","2E75B6"),("總公司","833C00")]:
        recs=[r for r in records if r['cat']==cat_name]
        if not recs: continue
        ws.row_dimensions[row].height=18
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=NCOL)
        c=ws.cell(row=row,column=1,value=f"◆ {cat_name}")
        c.font=FC; c.fill=fl(cf); c.alignment=cx("left"); c.border=BDR2
        row+=1; sub=0
        for i,r in enumerate(recs):
            ws.row_dimensions[row].height=18
            f2=fl("EBF0FA") if i%2==0 else fl("FFFFFF")
            vals=[r['store'],r['cat'],r['water_no'],r['pay_date'],r['period'],
                  r['invoice'],r['inv_ym'],r['carrier'],r['tax_id'],
                  r['tax'],r['fee'],r['total'],r['name'],r['address']]
            for c2,v in enumerate(vals,1):
                cl=wcell(ws,row,c2,v,FD,f2,col_align(c2),BDR)
                if c2 in{10,11,12}: cl.number_format='#,##0'
            sub+=r['total']; row+=1
        ws.row_dimensions[row].height=18
        for c2 in range(1,NCOL+1):
            if c2==1: wcell(ws,row,c2,f"{cat_name}小計",FT,fl("D9E1F2"),cx(),BDR)
            elif c2==12: wcell(ws,row,c2,sub,FT,fl("D9E1F2"),cx("right"),BDR,'#,##0')
            else: wcell(ws,row,c2,None,FT,fl("D9E1F2"),col_align(c2),BDR)
        grand+=sub; row+=2
    ws.row_dimensions[row].height=20
    for c2 in range(1,NCOL+1):
        if c2==1: wcell(ws,row,c2,"總計",FT,fl("D9E1F2"),cx(),BDR)
        elif c2==12: wcell(ws,row,c2,grand,FT,fl("D9E1F2"),cx("right"),BDR,'#,##0')
        else: wcell(ws,row,c2,None,FT,fl("D9E1F2"),col_align(c2),BDR)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def label_from_raw(raw):
    m=re.match(r'^\d+[-\s]*(.*)',str(raw).strip())
    if m:
        t=m.group(1).strip(); return t if t else str(raw)
    return str(raw)

def period_summary(p):
    if '-' in p: return p+"水費"
    s,e=p.split('/'); return f"{s[:3]}/{s[3:5]}/{s[5:]}-{e[:3]}/{e[3:5]}/{e[5:]}水費"

def make_xls(records, cat_name, pay_date):
    is_hq=(cat_name=="總公司")
    debit=622050 if is_hq else 612050; credit=110122001 if is_hq else 612050
    def bf(b=False): f=xlwt.Font(); f.name='Arial'; f.bold=b; f.height=200; return f
    def bd():
        b=xlwt.Borders()
        for s in('left','right','top','bottom'): setattr(b,s,xlwt.Borders.THIN)
        return b
    def fp(c): p=xlwt.Pattern(); p.pattern=xlwt.Pattern.SOLID_PATTERN; p.pattern_fore_colour=c; return p
    def sty(font,pat=None,b=None,h=xlwt.Alignment.HORZ_CENTER):
        s=xlwt.XFStyle(); s.font=font
        if pat: s.pattern=pat
        if b: s.borders=b
        a=xlwt.Alignment(); a.horz=h; a.vert=xlwt.Alignment.VERT_CENTER; s.alignment=a; return s
    b=bd(); fn=bf(); fnb=bf(True); L=xlwt.Alignment.HORZ_LEFT
    st={'hdr':sty(fnb,fp(0x1F),b),'c':sty(fn,None,b),'ca':sty(fn,fp(0x16),b),
        'l':sty(fn,None,b,L),'la':sty(fn,fp(0x16),b,L),
        'r':sty(fn,None,b,xlwt.Alignment.HORZ_RIGHT),
        'ra':sty(fn,fp(0x16),b,xlwt.Alignment.HORZ_RIGHT)}
    VHDRS=["項次","科目代號","部門","對象/廠商/客戶","**沖銷文件","忽略檢查~沖銷文件","幣別","原幣金額",
           "匯率","借方金額","貸方金額","摘要","發票未稅/~LC開狀金額","借款~起始日","到期日",
           "借款種類~利率%","還R/續E~廠商","文件說明","數量統計","單價","現金~流量","D/C","Comment"]
    VA=['c','c','c','c','c','c','c','c','c','r','r','l','c','c','c','c','c','l','c','c','c','c','l']
    VW=[6,10,10,12,12,14,6,10,6,10,10,30,14,10,10,14,10,10,8,8,10,6,10]
    wbx=xlwt.Workbook(encoding='utf-8'); ws2=wbx.add_sheet('工作表1')
    ws2.row(0).height_mismatch=True; ws2.row(0).height=400
    for c,h in enumerate(VHDRS): ws2.write(0,c,h,st['hdr'])
    for i,w in enumerate(VW): ws2.col(i).width=w*256
    total_credit=sum(r['total'] for r in records)
    vrows=[]; seq=1
    for r in records:
        vrows.append([seq,debit,r['store'],None,None,None,None,0,None,r['fee'],0,
                      period_summary(r['period']),None,None,None,None,None,None,None,None,None,"D",None])
        seq+=1
        tax_dept="001" if is_hq else r['store']
        vrows.append([seq,111020,tax_dept,"99999",f"{r['inv_ym']}99",None,None,0,None,r['tax'],0,
                      f"{r['store']}水費5%",None,None,None,None,None,None,None,None,None,"D",None])
        seq+=1
    if is_hq:
        lbl=label_from_raw(records[0].get('raw', records[0]['store']))
        csummary=f"{lbl}水費扣款"
    else:
        pts=pay_date.split('/'); csummary=f"{int(pts[1])}/{int(pts[2])}水費沖轉各門市"
    vrows.append([seq,credit,"001",None,None,None,None,0,None,0,total_credit,
                  csummary,None,None,None,None,None,None,None,None,None,"C",None])
    for ri,row in enumerate(vrows,1):
        ws2.row(ri).height_mismatch=True; ws2.row(ri).height=360; alt=(ri%2==0)
        for c,val in enumerate(row):
            if val is None: val=''
            a=VA[c]; key=a+('a' if alt else '')
            ws2.write(ri,c,val,st.get(key,st['ca' if alt else 'c']))
    buf=io.BytesIO(); wbx.save(buf); buf.seek(0); return buf

# ── 介面 ──────────────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    lookup_file = st.file_uploader("① 上傳水號對照.xlsx", type=['xlsx'])
with col2:
    pdf_files = st.file_uploader("② 上傳水費PDF（可多選）",
                                  type=['pdf'], accept_multiple_files=True)

if st.button("▶ 開始處理", type="primary",
             disabled=(not lookup_file or not pdf_files)):
    lookup, cat_map = load_lookup(lookup_file)
    records_by_date = {}
    errors = []
    progress = st.progress(0)
    status = st.empty()

    for idx, pdf in enumerate(pdf_files):
        status.text(f"處理中：{pdf.name}")
        rec = process_pdf(pdf.read(), pdf.name)
        if not rec:
            errors.append(f"⚠️ 無法辨識：{pdf.name}"); continue
        raw = lookup.get(rec['water_no'], 'NOT FOUND')
        if raw == 'NOT FOUND':
            errors.append(f"⚠️ 水號未找到：{rec['water_no']} ({pdf.name})")
            store, cat, label = '??', '門市', '??'
        else:
            store, cat, label = parse_store(raw, cat_map)
        rec.update(store=store, cat=cat, label=label, raw=raw)
        records_by_date.setdefault(rec['pay_date'], []).append(rec)
        progress.progress((idx+1)/len(pdf_files))

    status.empty(); progress.empty()

    if errors:
        for e in errors: st.warning(e)

    st.success(f"✅ 完成！共 {len(pdf_files)} 個PDF，{len(records_by_date)} 個繳費日期批次")

    for pay_date, recs in sorted(records_by_date.items()):
        prefix = pay_date.replace('/','')
        st.markdown(f"---\n### 📅 {pay_date}（{len(recs)} 筆）")

        # 明細
        for r in recs:
            st.write(f"  {'✅' if r['store']!='??' else '⚠️'} "
                     f"`{r['water_no']}` → **{r['store']}** ({r['cat']})　"
                     f"應繳 **{r['total']:,}** 元")

        buf = make_xlsx(recs)
        st.download_button(f"⬇ {prefix}水費明細.xlsx",
                           data=buf, file_name=f"{prefix}水費明細.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        for cat in ["門市","總公司"]:
            cat_recs=[r for r in recs if r['cat']==cat]
            if cat_recs:
                buf = make_xls(cat_recs, cat, pay_date)
                st.download_button(f"⬇ {prefix}水費傳票-{cat}.xls",
                                   data=buf, file_name=f"{prefix}水費傳票-{cat}.xls",
                                   mime="application/vnd.ms-excel")
